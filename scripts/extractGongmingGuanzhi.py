# -*- coding: utf-8 -*-
"""
从「个人详情」提取「功名」「官职」并回填。
- 功名：学历资格（庠生/贡生/进士/国学生等）
- 官职：职衔岗位（实职、散阶、勅赠衔、近现代职务）

默认不覆盖已有非空字段；若官职误填为纯功名（如邑庠生），则纠正到功名。

Usage:
  python scripts/extractGongmingGuanzhi.py
  python scripts/extractGongmingGuanzhi.py --write
"""
from __future__ import annotations

import argparse
import json
import re
import shutil
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

SOURCE = Path(r"D:\家谱\【古】罗氏家谱.xlsx")
REPORT = Path(__file__).resolve().parents[1] / "database" / "gongming_guanzhi_extract_report.json"

# 长词优先
GONGMING_TERMS = sorted(
    [
        "补国学生",
        "国学生",
        "太学生",
        "大学生",
        "监生",
        "邑庠生",
        "郡庠生",
        "郡廪生",
        "邑廪生",
        "廪生",
        "廩生",
        "增生",
        "附生",
        "文庠生",
        "武庠生",
        "庠生",
        "岁贡生",
        "恩贡生",
        "拔贡生",
        "副贡生",
        "优贡生",
        "岁贡",
        "恩贡",
        "思贡",
        "拔贡",
        "贡生",
        "武进士",
        "武举人",
        "武举",
        "进士",
        "举人",
        "解元",
        "会元",
        "状元",
        "秀才",
        "生员",
    ],
    key=len,
    reverse=True,
)

# 散阶 / 常见虚衔 → 官职
Sanjie_TERMS = sorted(
    [
        "怀远将军",
        "昭勇将军",
        "明威将军",
        "定远将军",
        "忠显校尉",
        "迪功郎",
        "儒林郎",
        "文林郎",
        "征仕郎",
        "登仕郎",
        "修职郎",
        "承德郎",
        "承事郎",
        "宣德郎",
        "朝列大夫",
        "中宪大夫",
        "中议大夫",
        "奉政大夫",
        "奉直大夫",
        "微仕郎",
        "徵仕郎",
    ],
    key=len,
    reverse=True,
)

GONGMING_SET = set(GONGMING_TERMS) | {"岁贡生", "恩贡生", "思贡生"}
FEMALE_TITLE = re.compile(r"^(孺人|安人|宜人|恭人|淑人|夫人)$")
STOP_CUT = re.compile(r"(?=生[于乾嘉道咸同光宣明清洪永正景成弘正嘉隆万泰天崇康雍乾]|配|继配|续配|元配|子|女|殁|葬|奉|乐恬|诏|现退休|现居)")


def norm_gm(term: str) -> str:
    mapping = {
        "岁贡": "岁贡生",
        "恩贡": "恩贡生",
        "思贡": "思贡生",
        "拔贡": "拔贡生",
        "廩生": "廪生",
        "大学生": "太学生",
    }
    return mapping.get(term, term)


def strip_annotation(text: str) -> str:
    """去掉批注与 Word 换行符，以免干扰提取。"""
    t = text or ""
    # 谱文后的长注（Word/_x000D_ 之后多为今人考证）只取正文首段
    t = t.split("_x000D_")[0]
    t = t.replace("\r", " ")
    t = re.sub(r"（[^）]*注[^）]*）", "", t)
    t = re.sub(r"\([^)]*注[^)]*\)", "", t)
    t = re.sub(r"阳先注[:：]?.*$", "", t)
    t = re.sub(r"也就是说.*$", "", t)
    return t


def extract_gongming(text: str) -> list[str]:
    text = strip_annotation(text)
    hits: list[str] = []
    for term in GONGMING_TERMS:
        for m in re.finditer(re.escape(term), text):
            start = m.start()
            after = text[m.end() : m.end() + 16]
            before = text[max(0, start - 12) : start]
            # 他人功名：进士XX赠扁/匾
            if term in ("进士", "举人", "翰林") and re.match(
                r".{0,6}(赠扁|赠匾|寿文|为之序)", after
            ):
                continue
            if term == "举人" and re.search(r"为之序", after):
                continue
            # 「禾川举人，肖鸾仪」类
            if term == "举人" and re.match(r"[，,]\s*\S{2,8}为", after):
                continue
            # 寿文中的翰林院
            if "翰林院" in text[start : start + 6]:
                continue
            if re.search(r"(赠扁|赠匾)$", before):
                continue
            hits.append(norm_gm(term))
    # 去重保序，且去掉被更长短语覆盖的短词
    out: list[str] = []
    for h in hits:
        if h in out:
            continue
        # 若已有更完整词则跳过短词
        if any(h != o and h in o for o in out):
            continue
        # 新词覆盖旧短词
        out = [o for o in out if not (o != h and o in h)]
        out.append(h)
    return out


TITLE_HINT = re.compile(
    r"长|员|官|使|郎|尉|将|军|守|备|总|司|知|县|州|府|簿|丞|谕|导|授|检|"
    r"御史|佥事|副使|同知|通判|经理|书记|主任|院长|站长|股长|营长|连长|排长|"
    r"宾|大夫|校尉|主簿|守备"
)


def clean_office(raw: str) -> str | None:
    t = raw.strip(" ，,。；;、）)")
    t = re.sub(r"^担任", "", t)
    t = STOP_CUT.split(t)[0].strip(" ，,。；;、）)")
    t = re.sub(r"(多年|现退休在家|退休)$", "", t).strip()
    if not t or len(t) < 2 or len(t) > 28:
        return None
    if FEMALE_TITLE.match(t):
        return None
    if re.search(r"[*]|匾|扁|寿文|序$|前茅|余卷|案首|刚直|纠劾|谢恩|入都", t):
        return None
    # 子女名误切
    if re.fullmatch(r"[现升恒完香]秀|[永]\S{1,2}", t):
        return None
    if t in GONGMING_SET or norm_gm(t) in GONGMING_SET:
        return None
    # 纯地名碎片（如「洲湖」）无职务词则丢弃
    if not TITLE_HINT.search(t) and len(t) <= 4:
        return None
    return t


def extract_guanzhi(text: str) -> list[str]:
    text = strip_annotation(text)
    offices: list[str] = []

    def add(item: str | None):
        if not item:
            return
        # 赠衔标注
        if item not in offices:
            offices.append(item)

    # 1) 散阶/虚衔直接命中
    for term in Sanjie_TERMS:
        if term in text:
            # 迪功郎致仕 / 勅授儒林郎
            add(term)

    # 2) 勅/敕/诰 + 赠/授/封
    for m in re.finditer(r"(勅|敕|诰)(赠|授|封)([^，。；;\n]{2,20})", text):
        verb = m.group(2)
        body = clean_office(m.group(3))
        if not body:
            continue
        if FEMALE_TITLE.match(body):
            continue
        if "匾" in body or "扁" in body:
            continue
        if verb == "赠":
            add(f"{body}(赠)")
        else:
            add(body)

    # 3) 任/升/拜/历任/累官/官至/担任/曾任
    for m in re.finditer(
        r"(?:先后)?(?:担任|历任|累官|官至|曾任|任|升|拜)([^，。；;\n]{2,36})", text
    ):
        chunk = m.group(1)
        if re.search(r"前茅|取余卷|入大学|入学", chunk):
            continue
        # 「洲湖、彭坊邮电支局局长」：前段为纯地名时合并；「站长、主任」并列则拆分
        if "、" in chunk:
            segs = [s.strip() for s in chunk.split("、") if s.strip()]
            if (
                len(segs) >= 2
                and not TITLE_HINT.search(segs[0])
                and TITLE_HINT.search(segs[-1])
            ):
                merged = "".join(segs)
                merged = re.split(r"多年", merged)[0]
                add(clean_office(merged))
                continue
        parts = re.split(r"[、，/]|以及|(?<![营])和(?!营)", chunk)
        for p in parts:
            p = clean_office(p)
            if p:
                add(p)

    # 4) 以功为X / 庆典X致仕
    for m in re.finditer(r"以功为([^，。；;\n]{2,18})", text):
        add(clean_office(m.group(1)))
    for m in re.finditer(r"([^，。；;\s]{2,12})致仕", text):
        cand = m.group(1)
        # 只取末尾衔名
        for term in Sanjie_TERMS:
            if cand.endswith(term):
                add(term)
                break

    # 5) 钦授/勅授 乡饮大宾…
    for m in re.finditer(r"(?:钦授|勅授|敕授)([^，。；;\n]{2,24})", text):
        add(clean_office(m.group(1)))

    # 去重：赠衔优先；长衔覆盖短衔
    out: list[str] = []
    for o in offices:
        base = o[:-3] if o.endswith("(赠)") else o
        if o.endswith("(赠)"):
            out = [x for x in out if x != base and x != o]
            out.append(o)
        elif f"{o}(赠)" in out:
            continue
        elif o not in out:
            out.append(o)
    # 散阶短衔被长衔后缀包含时去掉（…微仕郎 ⊃ 微仕郎）；不影响 连长/支部书记
    filtered: list[str] = []
    for o in out:
        if re.search(r"(郎|大夫|将军|校尉)$", o) and any(
            o != other and other.endswith(o) and len(other) - len(o) >= 4
            for other in out
        ):
            continue
        filtered.append(o)
    return filtered


def join_gm(items: list[str]) -> str:
    return "; ".join(items)


def join_gz(items: list[str]) -> str:
    return "; ".join(items)


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--write", action="store_true")
    parser.add_argument("--source", default=str(SOURCE))
    args = parser.parse_args()
    source = Path(args.source)

    wb = load_workbook(source)
    ws = wb.active
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    idx = {h: i + 1 for i, h in enumerate(headers) if h}
    for k in ("个人详情", "功名", "官职", "姓名"):
        if k not in idx:
            raise SystemExit(f"缺少列: {k}")

    changes = []
    skipped_keep = []

    for r in range(2, ws.max_row + 1):
        detail = str(ws.cell(r, idx["个人详情"]).value or "")
        if not detail.strip():
            continue
        old_gm = ws.cell(r, idx["功名"]).value
        old_gz = ws.cell(r, idx["官职"]).value
        old_gm_s = str(old_gm).strip() if old_gm not in (None, "") else ""
        old_gz_s = str(old_gz).strip() if old_gz not in (None, "") else ""

        eg = extract_gongming(detail)
        eo = extract_guanzhi(detail)

        new_gm = old_gm_s
        new_gz = old_gz_s
        actions = []

        # 纠正：官职里误填纯功名
        if old_gz_s and norm_gm(old_gz_s) in GONGMING_SET:
            if not new_gm:
                new_gm = norm_gm(old_gz_s)
                actions.append("move_gz_to_gm")
            new_gz = ""
            actions.append("clear_misplaced_gz")

        if not new_gm and eg:
            new_gm = join_gm(eg)
            actions.append("fill_gm")
        elif new_gm:
            skipped_keep.append({"row": r, "field": "功名", "value": new_gm})

        if not new_gz and eo:
            new_gz = join_gz(eo)
            actions.append("fill_gz")

        if new_gm == old_gm_s and new_gz == old_gz_s:
            continue

        changes.append(
            {
                "row": r,
                "name": ws.cell(r, idx["姓名"]).value,
                "old_gm": old_gm_s,
                "new_gm": new_gm,
                "old_gz": old_gz_s,
                "new_gz": new_gz,
                "actions": actions,
                "detail": detail[:120],
            }
        )
        if args.write:
            ws.cell(r, idx["功名"]).value = new_gm or None
            ws.cell(r, idx["官职"]).value = new_gz or None

    report = {
        "source": str(source),
        "write": args.write,
        "changed": len(changes),
        "fill_gm": sum(1 for c in changes if "fill_gm" in c["actions"] or "move_gz_to_gm" in c["actions"]),
        "fill_gz": sum(1 for c in changes if "fill_gz" in c["actions"]),
        "changes": changes,
    }
    REPORT.parent.mkdir(parents=True, exist_ok=True)
    REPORT.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    print(
        json.dumps(
            {
                "write": args.write,
                "changed": report["changed"],
                "fill_gm": report["fill_gm"],
                "fill_gz": report["fill_gz"],
                "report": str(REPORT),
            },
            ensure_ascii=False,
            indent=2,
        )
    )
    print("samples:")
    for c in changes[:20]:
        print(
            f"  {c['row']} {c['name']}: 功名[{c['old_gm']}]→[{c['new_gm']}] | 官职[{c['old_gz']}]→[{c['new_gz']}]"
        )

    # 对照金标自检（不改已有）
    print("gold-check extract:")
    gold = {
        32: ("邑庠生", None),
        37: ("进士", "南京山东道监察御史"),
        39: ("贡生", "浙江桐乡县主簿"),
        40: ("补国学生", "福建水师营守备"),
        33: (None, "南京山东道监察御史"),
        13: (None, "迪功郎"),
        3: (None, "泾源节度使"),
    }
    for r, (g, z) in gold.items():
        detail = str(ws.cell(r, idx["个人详情"]).value or "")
        eg = extract_gongming(detail)
        eo = extract_guanzhi(detail)
        ok_g = (g is None) or (g in eg) or any(g in x for x in eg)
        ok_z = (z is None) or any(z in x for x in eo)
        print(f"  row{r} gm={eg} gz={eo} gm_ok={ok_g} gz_ok={ok_z}")

    if not args.write:
        print("Dry-run only. Re-run with --write to save.")
        return

    stamp = datetime.now().strftime("%Y%m%d%H%M%S")
    backup = source.with_name(f"【古】罗氏家谱_功名官职_{stamp}.xlsx")
    wb.save(source)
    shutil.copy2(source, backup)
    print(f"Written: {source}")
    print(f"Backup: {backup}")


if __name__ == "__main__":
    main()
