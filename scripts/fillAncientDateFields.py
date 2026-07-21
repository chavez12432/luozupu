# -*- coding: utf-8 -*-
"""
根据「出生日期」「逝世日期」覆盖回填：
  出生朝代 / 出生干支 / 生肖
  逝世朝代 / 逝世干支

日期格式为 年.月.日（月日按谱面数字转中文，不做公历→农历换算）。
1921 年及以后：朝代字段直接写「YYYY年」。

Usage:
  python scripts/fillAncientDateFields.py
  python scripts/fillAncientDateFields.py --write
"""
from __future__ import annotations

import argparse
import json
import re
import shutil
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

SOURCE = Path(r"D:\家谱\【古】罗氏家谱.xlsx")
ERAS_PATH = Path(__file__).resolve().parents[1] / "database" / "dynasty_eras.json"
REPORT = Path(__file__).resolve().parents[1] / "database" / "ancient_date_fill_report.json"

TIAN_GAN = list("甲乙丙丁戊己庚辛壬癸")
DI_ZHI = list("子丑寅卯辰巳午未申酉戌亥")
ZODIAC = list("鼠牛虎兔龙蛇马羊猴鸡狗猪")

CN_DIGITS = "零一二三四五六七八九"

# 宋帝庙号（按 dynasty_eras.json 中的 emperor 名）
SONG_TEMPLE = {
    "赵匡胤": "宋太祖",
    "赵光义": "宋太宗",
    "赵恒": "宋真宗",
    "赵祯": "宋仁宗",
    "赵曙": "宋英宗",
    "赵顼": "宋神宗",
    "赵煦": "宋哲宗",
    "赵佶": "宋徽宗",
    "赵桓": "宋钦宗",
    "赵构": "宋高宗",
    "赵昚": "宋孝宗",
    "赵惇": "宋光宗",
    "赵扩": "宋宁宗",
    "赵昀": "宋理宗",
    "赵禥": "宋度宗",
    "赵㬎": "宋恭帝",
    "赵昰": "宋端宗",
    "赵昺": "宋卫王",
}

# 明帝庙号/常见尊称（用于「明朝X宗年号」类表述时备用；明清按用户样例主要用年号）
MING_TEMPLE = {
    "朱元璋": "明太祖",
    "朱允炆": "明惠帝",
    "朱棣": "明成祖",
    "朱高炽": "明仁宗",
    "朱瞻基": "明宣宗",
    "朱祁镇": "明英宗",
    "朱祁钰": "明代宗",
    "朱见深": "明宪宗",
    "朱祐樘": "明孝宗",
    "朱厚照": "明武宗",
    "朱厚熜": "明世宗",
    "朱载垕": "明穆宗",
    "朱翊钧": "明神宗",
    "朱常洛": "明光宗",
    "朱由校": "明熹宗",
    "朱由检": "明思宗",
}

MODERN_YEAR = 1921


def year_to_cn(n: int) -> str:
    """年号序数：1→元，2→二，10→十，11→十一，20→二十，21→二十一"""
    if n <= 0:
        return str(n)
    if n == 1:
        return "元"
    if n < 10:
        return CN_DIGITS[n]
    if n == 10:
        return "十"
    if n < 20:
        return "十" + CN_DIGITS[n - 10]
    if n < 100:
        tens, ones = divmod(n, 10)
        head = CN_DIGITS[tens] + "十"
        return head if ones == 0 else head + CN_DIGITS[ones]
    return str(n)


def month_to_cn(m: int) -> str:
    names = {
        1: "元月",
        2: "二月",
        3: "三月",
        4: "四月",
        5: "五月",
        6: "六月",
        7: "七月",
        8: "八月",
        9: "九月",
        10: "十月",
        11: "十一月",
        12: "十二月",
    }
    return names.get(m, f"{m}月")


def day_to_cn(d: int) -> str:
    if d < 1 or d > 30:
        return str(d)
    if d < 10:
        return "初" + CN_DIGITS[d]
    if d == 10:
        return "初十"
    if d < 20:
        return "十" + CN_DIGITS[d - 10]
    if d == 20:
        return "二十"
    if d < 30:
        return "廿" + CN_DIGITS[d - 20]
    return "三十"


def ganzhi_of_year(year: int) -> tuple[str, str]:
    gan = TIAN_GAN[(year - 4) % 10]
    zhi = DI_ZHI[(year - 4) % 12]
    zodiac = ZODIAC[(year - 4) % 12]
    return gan + zhi, zodiac


def parse_date(raw) -> tuple[int, int | None, int | None] | None:
    if raw is None:
        return None
    if isinstance(raw, (int, float)):
        y = int(raw)
        if 100 <= y <= 2100:
            return y, None, None
        return None
    s = str(raw).strip()
    if not s:
        return None
    m = re.fullmatch(r"(\d{3,4})\.(\d{1,2})\.(\d{1,2})", s)
    if m:
        return int(m.group(1)), int(m.group(2)), int(m.group(3))
    m = re.fullmatch(r"(\d{3,4})", s)
    if m:
        return int(m.group(1)), None, None
    m = re.fullmatch(r"(\d{3,4})[-/](\d{1,2})[-/](\d{1,2})", s)
    if m:
        return int(m.group(1)), int(m.group(2)), int(m.group(3))
    return None


def load_eras() -> list[dict]:
    return json.loads(ERAS_PATH.read_text(encoding="utf-8"))


def find_era(year: int, eras: list[dict]) -> dict | None:
    hits = [e for e in eras if e["startYear"] <= year <= e["endYear"]]
    if not hits:
        return None
    # 重叠时取 startYear 最大者（更贴近年号更替）
    # 宋元重叠（约 1260–1279）优先宋
    song = [e for e in hits if e["dynasty"] == "宋"]
    if song and year <= 1279:
        hits = song
    return max(hits, key=lambda e: e["startYear"])


def format_dynasty(year: int, eras: list[dict]) -> str:
    if year >= MODERN_YEAR:
        return f"{year}年"
    if 1912 <= year <= 1920:
        return f"民国{year_to_cn(year - 1911)}年"

    era = find_era(year, eras)
    if not era:
        return f"{year}年"

    ord_cn = year_to_cn(year - era["startYear"] + 1)
    dynasty = era["dynasty"]
    era_name = era["eraName"]
    emperor = era.get("emperor") or ""

    if dynasty == "宋":
        period = "北宋" if year < 1127 else "南宋"
        temple = SONG_TEMPLE.get(emperor, "")
        if temple:
            return f"{period}{temple}{era_name}{ord_cn}年"
        return f"{period}{era_name}{ord_cn}年"

    if dynasty == "清":
        return f"清朝{era_name}{ord_cn}年"

    if dynasty == "明":
        # 与清朝样例一致：朝代 + 年号 + 序年；庙号可省略
        return f"明朝{era_name}{ord_cn}年"

    if dynasty == "元":
        return f"元朝{era_name}{ord_cn}年"

    return f"{dynasty}{era_name}{ord_cn}年"


def format_ganzhi(year: int, month: int | None, day: int | None) -> str:
    gz, _ = ganzhi_of_year(year)
    parts = [f"{gz}年"]
    if month is not None:
        parts.append(month_to_cn(month))
    if day is not None:
        parts.append(day_to_cn(day))
    return "".join(parts)


def convert_birth(raw, eras: list[dict]) -> dict | None:
    parsed = parse_date(raw)
    if not parsed:
        return None
    year, month, day = parsed
    gz, zodiac = ganzhi_of_year(year)
    return {
        "dynasty": format_dynasty(year, eras),
        "ganzhi": format_ganzhi(year, month, day),
        "zodiac": zodiac,
        "year": year,
    }


def convert_death(raw, eras: list[dict]) -> dict | None:
    parsed = parse_date(raw)
    if not parsed:
        return None
    year, month, day = parsed
    return {
        "dynasty": format_dynasty(year, eras),
        "ganzhi": format_ganzhi(year, month, day),
        "year": year,
    }


def self_check(eras: list[dict]) -> None:
    cases = [
        ("1104.1.1", "北宋宋徽宗崇宁三年", "甲申年元月初一", "猴"),
        ("1741.5.25", "清朝乾隆六年", "辛酉年五月廿五", "鸡"),
        ("1965.8.10", "1965年", "乙巳年八月初十", "蛇"),
        ("1994.11.6", "1994年", "甲戌年十一月初六", None),
    ]
    print("self-check:")
    for raw, exp_d, exp_g, exp_z in cases:
        if exp_z is not None:
            r = convert_birth(raw, eras)
            ok = r and r["dynasty"] == exp_d and r["ganzhi"] == exp_g and r["zodiac"] == exp_z
            print(f"  birth {raw} -> {r} {'OK' if ok else 'FAIL'}")
            if not ok:
                print(f"    expect dynasty={exp_d} ganzhi={exp_g} zodiac={exp_z}")
        else:
            r = convert_death(raw, eras)
            ok = r and r["dynasty"] == exp_d and r["ganzhi"] == exp_g
            print(f"  death {raw} -> {r} {'OK' if ok else 'FAIL'}")
            if not ok:
                print(f"    expect dynasty={exp_d} ganzhi={exp_g}")


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--write", action="store_true")
    parser.add_argument("--source", default=str(SOURCE))
    args = parser.parse_args()
    source = Path(args.source)

    if not source.exists():
        raise SystemExit(f"文件不存在: {source}")
    magic = source.read_bytes()[:8]
    if magic.startswith(b"%TSD"):
        raise SystemExit("文件仍是腾讯文档加密，请先另存为普通 xlsx")

    eras = load_eras()
    self_check(eras)

    wb = load_workbook(source)
    ws = wb.active
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    idx = {h: i + 1 for i, h in enumerate(headers) if h}
    required = ["出生日期", "出生朝代", "出生干支", "生肖", "逝世日期", "逝世朝代", "逝世干支"]
    missing = [k for k in required if k not in idx]
    if missing:
        raise SystemExit(f"缺少列: {missing}")

    changes = []
    skipped = []

    for r in range(2, ws.max_row + 1):
        birth_raw = ws.cell(r, idx["出生日期"]).value
        death_raw = ws.cell(r, idx["逝世日期"]).value
        row_change = {"row": r, "name": ws.cell(r, idx.get("姓名", 2)).value}

        birth = convert_birth(birth_raw, eras)
        if birth:
            row_change["birth"] = {
                "raw": str(birth_raw),
                "dynasty": birth["dynasty"],
                "ganzhi": birth["ganzhi"],
                "zodiac": birth["zodiac"],
            }
            if args.write:
                ws.cell(r, idx["出生朝代"]).value = birth["dynasty"]
                ws.cell(r, idx["出生干支"]).value = birth["ganzhi"]
                ws.cell(r, idx["生肖"]).value = birth["zodiac"]
        elif birth_raw not in (None, ""):
            skipped.append({"row": r, "field": "出生日期", "raw": str(birth_raw)})

        death = convert_death(death_raw, eras)
        if death:
            row_change["death"] = {
                "raw": str(death_raw),
                "dynasty": death["dynasty"],
                "ganzhi": death["ganzhi"],
            }
            if args.write:
                ws.cell(r, idx["逝世朝代"]).value = death["dynasty"]
                ws.cell(r, idx["逝世干支"]).value = death["ganzhi"]
        elif death_raw not in (None, ""):
            skipped.append({"row": r, "field": "逝世日期", "raw": str(death_raw)})

        if "birth" in row_change or "death" in row_change:
            changes.append(row_change)

    report = {
        "source": str(source),
        "write": args.write,
        "changedRows": len(changes),
        "birthFilled": sum(1 for c in changes if "birth" in c),
        "deathFilled": sum(1 for c in changes if "death" in c),
        "skipped": skipped,
        "samples": changes[:40],
    }
    REPORT.parent.mkdir(parents=True, exist_ok=True)
    REPORT.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")

    print(
        json.dumps(
            {
                "source": str(source),
                "write": args.write,
                "changedRows": report["changedRows"],
                "birthFilled": report["birthFilled"],
                "deathFilled": report["deathFilled"],
                "skipped": len(skipped),
                "report": str(REPORT),
            },
            ensure_ascii=False,
            indent=2,
        )
    )

    if not args.write:
        print("Dry-run only. Re-run with --write to save.")
        return

    stamp = datetime.now().strftime("%Y%m%d%H%M%S")
    backup = source.with_name(f"【古】罗氏家谱_日期回填_{stamp}.xlsx")
    wb.save(source)
    shutil.copy2(source, backup)
    print(f"Written: {source}")
    print(f"Backup: {backup}")


if __name__ == "__main__":
    main()
