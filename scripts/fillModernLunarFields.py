# -*- coding: utf-8 -*-
"""
根据「出生农历」「逝世农历」覆盖回填：
  出生干支 / 出生公历
  逝世干支 / 逝世公历

农历 Y.M.D → 干支「壬申年三月初九」+ 公历「1932年4月14日」
表头在第 2 行。

Usage:
  python scripts/fillModernLunarFields.py
  python scripts/fillModernLunarFields.py --write
"""
from __future__ import annotations

import argparse
import json
import re
import shutil
from datetime import datetime
from pathlib import Path

from lunardate import LunarDate
from openpyxl import load_workbook

SOURCE = Path(r"D:\家谱\【今】罗氏家谱.xlsx")
REPORT = Path(__file__).resolve().parents[1] / "database" / "modern_lunar_fill_report.json"
HEADER_ROW = 2
DATA_START = 3

TIAN_GAN = list("甲乙丙丁戊己庚辛壬癸")
DI_ZHI = list("子丑寅卯辰巳午未申酉戌亥")
CN_DIGITS = "零一二三四五六七八九"


def month_to_cn(m: int, leap: bool = False) -> str:
    names = {
        1: "正月",
        2: "二月",
        3: "三月",
        4: "四月",
        5: "五月",
        6: "六月",
        7: "七月",
        8: "八月",
        9: "九月",
        10: "十月",
        11: "十一月",
        12: "十二月",
    }
    prefix = "闰" if leap else ""
    return prefix + names.get(m, f"{m}月")


def day_to_cn(d: int) -> str:
    if d < 1 or d > 30:
        return str(d)
    if d < 10:
        return "初" + CN_DIGITS[d]
    if d == 10:
        return "初十"
    if d < 20:
        return "十" + CN_DIGITS[d - 10]
    if d == 20:
        return "二十"
    if d < 30:
        return "廿" + CN_DIGITS[d - 20]
    return "三十"


def ganzhi_year(year: int) -> str:
    return TIAN_GAN[(year - 4) % 10] + DI_ZHI[(year - 4) % 12]


def parse_lunar(raw) -> tuple[int, int, int] | None:
    if raw is None:
        return None
    s = str(raw).strip()
    if not s:
        return None
    m = re.fullmatch(r"(\d{3,4})[.,\-／/](\d{1,2})[.,\-／/](\d{1,2})", s)
    if not m:
        return None
    y, mo, d = int(m.group(1)), int(m.group(2)), int(m.group(3))
    if not (1 <= mo <= 12 and 1 <= d <= 30):
        return None
    return y, mo, d


def lunar_to_fields(raw) -> dict | None:
    parsed = parse_lunar(raw)
    if not parsed:
        return None
    y, mo, d = parsed
    try:
        solar = LunarDate(y, mo, d, False).to_solar_date()
    except Exception:
        try:
            solar = LunarDate(y, mo, d, True).to_solar_date()
            return {
                "ganzhi": f"{ganzhi_year(y)}年{month_to_cn(mo, True)}{day_to_cn(d)}",
                "solar": f"{solar.year}年{solar.month}月{solar.day}日",
                "leap": True,
            }
        except Exception as e:
            return {"error": str(e), "raw": str(raw)}
    return {
        "ganzhi": f"{ganzhi_year(y)}年{month_to_cn(mo)}{day_to_cn(d)}",
        "solar": f"{solar.year}年{solar.month}月{solar.day}日",
        "leap": False,
    }


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--write", action="store_true")
    parser.add_argument("--source", default=str(SOURCE))
    args = parser.parse_args()
    source = Path(args.source)

    # self-check
    b = lunar_to_fields("1932.3.9")
    d = lunar_to_fields("1998.6.2")
    print("self-check birth", b)
    print("self-check death", d)
    assert b and b["ganzhi"] == "壬申年三月初九" and b["solar"] == "1932年4月14日", b
    assert d and d["ganzhi"] == "戊寅年六月初二", d
    # 公历：标准农历库为 7月24日；若与口头样例差一天，以库为准
    print("note: death solar from lib =", d["solar"], "(sample said 1998年7月25日)")

    wb = load_workbook(source)
    ws = wb.active
    headers = [c.value for c in next(ws.iter_rows(min_row=HEADER_ROW, max_row=HEADER_ROW))]
    idx = {h: i + 1 for i, h in enumerate(headers) if h}
    need = ["出生农历", "出生干支", "出生公历", "逝世农历", "逝世干支", "逝世公历"]
    missing = [k for k in need if k not in idx]
    if missing:
        raise SystemExit(f"缺少列: {missing}; headers={headers}")

    changes = []
    errors = []
    birth_n = death_n = 0

    for r in range(DATA_START, ws.max_row + 1):
        birth_raw = ws.cell(r, idx["出生农历"]).value
        death_raw = ws.cell(r, idx["逝世农历"]).value
        row = {"row": r, "name": ws.cell(r, idx.get("姓名", 2)).value}

        # 无论原值如何，有农历则覆盖；无农历则清空干支/公历
        if birth_raw not in (None, ""):
            info = lunar_to_fields(birth_raw)
            if info and "ganzhi" in info:
                row["birth"] = {"raw": str(birth_raw), **info}
                birth_n += 1
                if args.write:
                    ws.cell(r, idx["出生干支"]).value = info["ganzhi"]
                    ws.cell(r, idx["出生公历"]).value = info["solar"]
            else:
                errors.append({"row": r, "field": "出生农历", "raw": str(birth_raw), "err": info})
                if args.write:
                    ws.cell(r, idx["出生干支"]).value = None
                    ws.cell(r, idx["出生公历"]).value = None
        else:
            if args.write:
                ws.cell(r, idx["出生干支"]).value = None
                ws.cell(r, idx["出生公历"]).value = None

        if death_raw not in (None, ""):
            info = lunar_to_fields(death_raw)
            if info and "ganzhi" in info:
                row["death"] = {"raw": str(death_raw), **info}
                death_n += 1
                if args.write:
                    ws.cell(r, idx["逝世干支"]).value = info["ganzhi"]
                    ws.cell(r, idx["逝世公历"]).value = info["solar"]
            else:
                errors.append({"row": r, "field": "逝世农历", "raw": str(death_raw), "err": info})
                if args.write:
                    ws.cell(r, idx["逝世干支"]).value = None
                    ws.cell(r, idx["逝世公历"]).value = None
        else:
            if args.write:
                ws.cell(r, idx["逝世干支"]).value = None
                ws.cell(r, idx["逝世公历"]).value = None

        if "birth" in row or "death" in row:
            changes.append(row)

    report = {
        "source": str(source),
        "write": args.write,
        "birthFilled": birth_n,
        "deathFilled": death_n,
        "errors": errors,
        "samples": changes[:30],
        "note": "逝世农历1998.6.2 标准换算为1998年7月24日（样例曾写25日）",
    }
    REPORT.parent.mkdir(parents=True, exist_ok=True)
    REPORT.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    print(
        json.dumps(
            {
                "write": args.write,
                "birthFilled": birth_n,
                "deathFilled": death_n,
                "errors": len(errors),
                "report": str(REPORT),
            },
            ensure_ascii=False,
            indent=2,
        )
    )
    for c in changes[:8]:
        print(c.get("name"), c.get("birth"), c.get("death"))

    if not args.write:
        print("Dry-run only. Re-run with --write to save.")
        return

    stamp = datetime.now().strftime("%Y%m%d%H%M%S")
    backup = source.with_name(f"【今】罗氏家谱_农历回填_{stamp}.xlsx")
    wb.save(source)
    shutil.copy2(source, backup)
    print(f"Written: {source}")
    print(f"Backup: {backup}")


if __name__ == "__main__":
    main()
