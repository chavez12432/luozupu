# -*- coding: utf-8 -*-
"""
【今】罗氏家谱：
1) 由「出生公历」回填「星座」（覆盖）
2) 由「出生干支」插入生肖，改为「甲子鼠年三月初一」格式（覆盖）

Usage:
  python scripts/fillModernZodiacAndShengxiao.py
  python scripts/fillModernZodiacAndShengxiao.py --write
"""
from __future__ import annotations

import argparse
import json
import re
import shutil
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

SOURCE = Path(r"D:\家谱\【今】罗氏家谱.xlsx")
REPORT = Path(__file__).resolve().parents[1] / "database" / "modern_zodiac_shengxiao_report.json"
HEADER_ROW = 2
DATA_START = 3

ZHI_ANIMAL = {
    "子": "鼠",
    "丑": "牛",
    "寅": "虎",
    "卯": "兔",
    "辰": "龙",
    "巳": "蛇",
    "午": "马",
    "未": "羊",
    "申": "猴",
    "酉": "鸡",
    "戌": "狗",
    "亥": "猪",
}

# 星座分界（含起始月日）
CONSTELLATION_BOUNDARIES = [
    ((1, 20), "水瓶座"),
    ((2, 19), "双鱼座"),
    ((3, 21), "白羊座"),
    ((4, 20), "金牛座"),
    ((5, 21), "双子座"),
    ((6, 22), "巨蟹座"),
    ((7, 23), "狮子座"),
    ((8, 23), "处女座"),
    ((9, 23), "天秤座"),
    ((10, 24), "天蝎座"),
    ((11, 23), "射手座"),
    ((12, 22), "摩羯座"),
]


def parse_solar(raw):
    if raw is None:
        return None
    s = str(raw).strip()
    if not s:
        return None
    m = re.fullmatch(r"(\d{4})年(\d{1,2})月(\d{1,2})日", s)
    if m:
        return int(m.group(1)), int(m.group(2)), int(m.group(3))
    m = re.fullmatch(r"(\d{4})[./\-](\d{1,2})[./\-](\d{1,2})", s)
    if m:
        return int(m.group(1)), int(m.group(2)), int(m.group(3))
    return None


def constellation_of(month: int, day: int) -> str:
    md = (month, day)
    if md < (1, 20):
        return "摩羯座"
    result = "摩羯座"
    for start, name in CONSTELLATION_BOUNDARIES:
        if md >= start:
            result = name
        else:
            break
    return result


def insert_shengxiao(ganzhi_text: str) -> str | None:
    """甲子年三月初一 / 甲子年 三月初一 → 甲子鼠年三月初一"""
    if not ganzhi_text:
        return None
    s = str(ganzhi_text).strip()
    # already has animal
    if re.search(
        r"[甲乙丙丁戊己庚辛壬癸][子丑寅卯辰巳午未申酉戌亥][鼠牛虎兔龙蛇马羊猴鸡狗猪]年",
        s,
    ):
        return re.sub(r"\s+", "", s)
    m = re.match(
        r"^([甲乙丙丁戊己庚辛壬癸][子丑寅卯辰巳午未申酉戌亥])\s*年\s*(.*)$",
        s,
    )
    if not m:
        return None
    gz, rest = m.group(1), m.group(2).strip()
    animal = ZHI_ANIMAL[gz[1]]
    return f"{gz}{animal}年{rest}" if rest else f"{gz}{animal}年"


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--write", action="store_true")
    parser.add_argument("--source", default=str(SOURCE))
    args = parser.parse_args()
    source = Path(args.source)

    assert constellation_of(4, 1) == "白羊座"
    assert insert_shengxiao("甲子年 三月初一") == "甲子鼠年三月初一"
    assert insert_shengxiao("壬申年三月初九") == "壬申猴年三月初九"
    print("self-check OK: 1984-4-1→白羊座, 甲子年→甲子鼠年, 壬申年→壬申猴年")

    wb = load_workbook(source)
    ws = wb.active
    headers = [c.value for c in next(ws.iter_rows(min_row=HEADER_ROW, max_row=HEADER_ROW))]
    idx = {h: i + 1 for i, h in enumerate(headers) if h}
    for k in ("出生公历", "星座", "出生干支"):
        if k not in idx:
            raise SystemExit(f"缺少列: {k}")

    star_n = shengxiao_n = 0
    changes = []
    skipped = []

    for r in range(DATA_START, ws.max_row + 1):
        solar_raw = ws.cell(r, idx["出生公历"]).value
        gz_raw = ws.cell(r, idx["出生干支"]).value
        name = ws.cell(r, idx["姓名"]).value
        row_chg = {"row": r, "name": name}

        solar = parse_solar(solar_raw)
        if solar:
            star = constellation_of(solar[1], solar[2])
            row_chg["星座"] = star
            star_n += 1
            if args.write:
                ws.cell(r, idx["星座"]).value = star
        else:
            if args.write:
                ws.cell(r, idx["星座"]).value = None
            if solar_raw not in (None, ""):
                skipped.append({"row": r, "field": "出生公历", "raw": str(solar_raw)})

        new_gz = insert_shengxiao(str(gz_raw).strip() if gz_raw else "")
        if new_gz:
            row_chg["出生干支"] = {"old": str(gz_raw) if gz_raw else "", "new": new_gz}
            shengxiao_n += 1
            if args.write:
                ws.cell(r, idx["出生干支"]).value = new_gz
        else:
            if args.write and gz_raw not in (None, ""):
                # 无法解析则清空，按“清除原有再回填”口径
                pass
            if gz_raw not in (None, "") and not new_gz:
                skipped.append({"row": r, "field": "出生干支", "raw": str(gz_raw)})

        if "星座" in row_chg or "出生干支" in row_chg:
            changes.append(row_chg)

    report = {
        "source": str(source),
        "write": args.write,
        "starFilled": star_n,
        "shengxiaoFilled": shengxiao_n,
        "skipped": skipped,
        "samples": changes[:40],
    }
    REPORT.parent.mkdir(parents=True, exist_ok=True)
    REPORT.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    print(
        json.dumps(
            {
                "write": args.write,
                "starFilled": star_n,
                "shengxiaoFilled": shengxiao_n,
                "skipped": len(skipped),
                "report": str(REPORT),
            },
            ensure_ascii=False,
            indent=2,
        )
    )
    for c in changes[:10]:
        print(c)

    if not args.write:
        print("Dry-run only. Re-run with --write to save.")
        return

    stamp = datetime.now().strftime("%Y%m%d%H%M%S")
    backup = source.with_name(f"【今】罗氏家谱_星座生肖_{stamp}.xlsx")
    wb.save(source)
    shutil.copy2(source, backup)
    print(f"Written: {source}")
    print(f"Backup: {backup}")


if __name__ == "__main__":
    main()
