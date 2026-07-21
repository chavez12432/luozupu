# -*- coding: utf-8 -*-
"""
按交叉校验清单修正 D:\\家谱\\database 四张 Excel 的阻塞性错误。

Usage:
  python scripts/fixExcelFamilyDb.py
  python scripts/fixExcelFamilyDb.py --dry-run
"""
from __future__ import annotations

import argparse
import json
import re
import shutil
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

SRC = Path(r"D:\家谱\database")
REPORT = Path(__file__).resolve().parents[1] / "database" / "excel_fix_report.json"

ANCIENT = SRC / "【古】罗氏家谱.xlsx"
MODERN = SRC / "【今】罗氏家谱.xlsx"
WIVES = SRC / "妻子表.xlsx"
SIL = SRC / "女婿表.xlsx"


def clean_text(v):
    if v is None:
        return None
    s = str(v).replace("_x000D_", "").replace("\r", "").replace("\n", " ").strip()
    return s or None


def find_header(ws, max_scan=5):
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=max_scan, values_only=False), start=1):
        vals = [str(c.value).strip() if c.value is not None else "" for c in row]
        if "ID" in vals:
            return i, {h: j for j, h in enumerate(vals) if h}, vals
    raise RuntimeError("未找到含 ID 的表头")


def cell(ws, row, col_idx):
    return ws.cell(row=row, column=col_idx + 1)


def find_row(ws, header_row, colmap, member_id):
    id_col = colmap["ID"]
    for r in range(header_row + 1, ws.max_row + 1):
        v = cell(ws, r, id_col).value
        if v is not None and str(v).strip() == member_id:
            return r
    return None


def set_field(ws, row, colmap, field, value, changes, file_label, member_id):
    if field not in colmap:
        return
    c = cell(ws, row, colmap[field])
    old = c.value
    if old == value:
        return
    c.value = value
    changes.append({
        "file": file_label,
        "id": member_id,
        "field": field,
        "old": old,
        "new": value,
    })


def append_spouse_id(existing, add_id):
    parts = []
    if existing is not None and str(existing).strip():
        parts = [p.strip() for p in re.split(r"[,，;；、\s]+", str(existing)) if p.strip()]
    if add_id not in parts:
        parts.append(add_id)
    return ",".join(parts)


def fix_members_sheet(path, file_label, fixes, changes, dry_run):
    wb = load_workbook(path)
    ws = wb.active
    header_row, colmap, _ = find_header(ws)
    for member_id, fields in fixes.items():
        row = find_row(ws, header_row, colmap, member_id)
        if row is None:
            changes.append({"file": file_label, "id": member_id, "error": "行未找到"})
            continue
        for field, value in fields.items():
            if field == "配偶ID_append":
                c = cell(ws, row, colmap["配偶ID"])
                new_v = append_spouse_id(c.value, value)
                set_field(ws, row, colmap, "配偶ID", new_v, changes, file_label, member_id)
            else:
                set_field(ws, row, colmap, field, value, changes, file_label, member_id)
    if not dry_run:
        backup = path.with_suffix(path.suffix + f".bak_{datetime.now().strftime('%Y%m%d_%H%M%S')}")
        shutil.copy2(path, backup)
        wb.save(path)
        changes.append({"file": file_label, "backup": str(backup)})
    wb.close()


def add_missing_wives(changes, dry_run):
    wb = load_workbook(WIVES)
    ws = wb["妻子表"] if "妻子表" in wb.sheetnames else wb.active
    header_row, colmap, headers = find_header(ws)
    existing = set()
    id_col = colmap["ID"]
    for r in range(header_row + 1, ws.max_row + 1):
        v = cell(ws, r, id_col).value
        if v is not None:
            existing.add(str(v).strip())

    to_add = [
        {
            "ID": "M002038W01",
            "丈夫ID": "M002038",
            "丈夫姓名": "闵鉴",
            "姓名": "（佚名）",
            "家乡": None,
            "世代": 19,
            "婚配类型": "配",
            "婚配序号": 1,
            "是否本村人": "否",
        },
        {
            "ID": "M000050W02",
            "丈夫ID": "M000050",
            "丈夫姓名": "守诚",
            "姓名": "（佚名）",
            "家乡": None,
            "世代": 23,
            "婚配类型": "继配",
            "婚配序号": 2,
            "是否本村人": "否",
        },
    ]
    for row_data in to_add:
        wid = row_data["ID"]
        if wid in existing:
            continue
        if dry_run:
            changes.append({"file": "妻子表", "id": wid, "action": "would_add", "data": row_data})
            continue
        new_row = ws.max_row + 1
        for h in headers:
            if not h:
                continue
            if h in row_data:
                cell(ws, new_row, colmap[h]).value = row_data[h]
        changes.append({"file": "妻子表", "id": wid, "action": "added", "data": row_data})

    # 修正异常姓名「妻生子」→ 保留但标注；不改名以免臆造，仅记录
    row = find_row(ws, header_row, colmap, "M000139W02")
    if row:
        name = cell(ws, row, colmap["姓名"]).value
        if name == "妻生子":
            changes.append({
                "file": "妻子表",
                "id": "M000139W02",
                "warning": "姓名为「妻生子」，疑似脏数据，未自动改名",
            })

    if not dry_run:
        backup = WIVES.with_suffix(WIVES.suffix + f".bak_{datetime.now().strftime('%Y%m%d_%H%M%S')}")
        shutil.copy2(WIVES, backup)
        wb.save(WIVES)
        changes.append({"file": "妻子表", "backup": str(backup)})
    wb.close()


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()
    changes = []

    ancient_fixes = {
        "M000043": {"父亲姓名": "应真"},
        "M000044": {"父亲姓名": "宏道"},
        "M000045": {"父亲姓名": "宏道"},
        "M001000": {"性别": "男"},
        "ZA0220": {"性别": "男"},
        "ZA0226": {"女婿ID": "ZA0226S01", "女婿姓名": "张家里"},
        "M000139": {"配偶ID_append": "M000139W02"},
        "M000170": {"配偶ID_append": "M000170W02"},
        "M000187": {"配偶ID_append": "M000187W02"},
        "M001228": {"配偶ID_append": "M001228W02"},
    }
    modern_fixes = {
        "M001258": {"性别": "男"},
        "M001317": {"性别": "男"},
        "M000490": {"女婿姓名": "黄文高"},
    }

    fix_members_sheet(ANCIENT, "古谱", ancient_fixes, changes, args.dry_run)
    fix_members_sheet(MODERN, "今谱", modern_fixes, changes, args.dry_run)
    add_missing_wives(changes, args.dry_run)

    report = {
        "dry_run": args.dry_run,
        "time": datetime.now().isoformat(timespec="seconds"),
        "changes": changes,
        "change_count": len([c for c in changes if "backup" not in c and "warning" not in c]),
    }
    REPORT.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Wrote {REPORT}")
    print(f"changes={report['change_count']} dry_run={args.dry_run}")


if __name__ == "__main__":
    main()
