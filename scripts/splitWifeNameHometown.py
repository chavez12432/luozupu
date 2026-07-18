# -*- coding: utf-8 -*-
"""
从妻子表「姓名」中拆分家乡前缀。
规则 A：家乡 + 姓氏氏 + 名
规则 B：家乡 + 现代姓名（无氏）
不符合则跳过。保留「氏」。家乡已有值时不覆盖，只改姓名。

Usage:
  python scripts/splitWifeNameHometown.py
  python scripts/splitWifeNameHometown.py --write
"""
from __future__ import annotations

import argparse
import json
import re
import shutil
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

SOURCE = Path(r"D:\家谱\database\妻子表.xlsx")
REPORT = Path(__file__).resolve().parents[1] / "database" / "wife_name_hometown_split_report.json"

SURNAMES = set(
    "赵钱孙李周吴郑王冯陈褚卫蒋沈韩杨朱秦尤许何吕施张孔曹严华金魏陶姜戚谢邹喻柏水窦章云苏潘葛奚范彭郎鲁韦昌马苗凤花方俞任袁柳鲍史唐费廉岑薛雷贺倪汤滕殷罗毕郝邬安常乐于时傅皮卞齐康伍余元卜顾孟平黄和穆萧肖尹姚邵湛汪祁毛禹狄米贝明臧计伏成戴谈宋茅庞熊纪舒屈项祝董梁杜阮蓝闵席季麻强贾路娄危江童颜郭梅盛林刁钟徐邱骆高夏蔡田樊胡凌霍虞万支柯管卢莫经房裘缪干解应宗丁宣邓郁单杭洪包诸左石崔吉钮龚程嵇邢滑裴陆荣翁荀羊於惠甄曲家封储靳井段富巫乌焦巴弓牧隗山谷车侯宓蓬全郗班仰秋仲伊宫宁仇栾暴甘钭厉戎祖武符刘景詹束龙叶幸司韶郜黎蓟薄印宿白怀蒲邰从鄂索咸籍赖卓蔺屠蒙池乔阴胥能苍双闻莘党翟谭贡劳逄姬申扶堵冉宰郦雍桑桂濮牛寿通边扈燕冀浦尚农温别庄晏柴瞿阎充慕连茹习宦艾鱼容向古易慎戈廖庾终暨居衡步都耿满弘匡国文寇广禄阙东欧殳沃利蔚越夔隆师巩聂晁勾敖融冷訾辛阚那简饶空曾毋沙乜养鞠须丰巢关蒯相查后荆红游竺权逯盖益桓公旷岩贺"
)
COMPOUND = ["欧阳", "司马", "上官", "诸葛", "司徒", "皇甫", "夏侯", "尉迟"]

GEO_HAS = re.compile(
    r"村|县|乡|镇|市|洲|溪|城|田|江|湖|头|里|门|坂|园|山|塘|桥|铺|坊|垅|堡|寨|坪|湾|窝|岭|岗|坑|源|社|"
    r"本土|本里|本村|本户|上城|前溪|永新|沧洲|槎江|沛溪|金田|彭坊|洲湖|下沧|堌头|罗坂|安福|萍乡|洋门|"
    r"烟头|中狮|承口|虹桥|樟木|城江|田心|社上|贵州|新余|桂林"
)
GEO_TAIL = re.compile(
    r"(村|县|乡|镇|市|洲|溪|城|坊|垅|里|头|田|江|湖|山|塘|桥|铺|堡|寨|坪|湾|岭|岗|坑|源|坂|门|园)$"
)
GEO_IN_GIVEN = re.compile(r"[村县乡镇市洲溪城田江湖头里门坂园山塘桥铺坊垅堡寨坪湾岭岗坑源社]")
PLACE_ONLY_END = re.compile(r"(乡|县|村|镇|市|省|路|街)$")
BAD = re.compile(r"殁|生年|配|继娶|妻子|地质|轻工|岁|研")
PURE_SHI = re.compile(r"^[\u4e00-\u9fff]{1,2}氏[\u4e00-\u9fff]{0,3}$")


def _surname_and_given_modern(rest: str):
    for c in COMPOUND:
        if rest.startswith(c):
            return c, rest[len(c) :]
    if rest and rest[0] in SURNAMES:
        return rest[0], rest[1:]
    return None, None


def is_person_shi(rest: str) -> bool:
    m = re.fullmatch(r"([\u4e00-\u9fff]{1,2})氏([\u4e00-\u9fff]{0,4})", rest)
    if not m:
        return False
    sur = m.group(1)
    if sur in ("某",) or sur in SURNAMES:
        return True
    return any(rest.startswith(c + "氏") for c in COMPOUND)


def is_person_modern(rest: str) -> bool:
    if not rest or "氏" in rest:
        return False
    sur, given = _surname_and_given_modern(rest)
    if sur is None:
        return False
    if not (1 <= len(given) <= 3):
        return False
    if PLACE_ONLY_END.search(given) or GEO_IN_GIVEN.search(given):
        return False
    return True


def place_ok(place: str) -> bool:
    if len(place) < 2:
        return False
    if not GEO_HAS.search(place):
        return False
    # 防吞姓：家乡以姓氏字结尾且非地名尾缀
    if place[-1] in SURNAMES and not GEO_TAIL.search(place):
        return False
    if len(place) >= 2 and place[-2:] in COMPOUND:
        return False
    return True


def try_split(name: str):
    name = (name or "").strip()
    if not name or BAD.search(name):
        return None
    # 无家乡前缀的纯 X氏 / X氏某
    if PURE_SHI.fullmatch(name):
        return None
    if len(name) <= 3 and "氏" not in name:
        return None

    if "氏" in name:
        for i in range(len(name) - 2, 1, -1):
            place, rest = name[:i], name[i:]
            if is_person_shi(rest) and len(place) >= 2:
                return {"hometown": place, "name": rest, "rule": "shi"}
        return None

    # 现代名：收集候选，优先「名中无地理字、人名较短、家乡有尾缀」
    candidates = []
    for i in range(2, len(name) - 1):
        place, rest = name[:i], name[i:]
        if not is_person_modern(rest):
            continue
        sur, given = _surname_and_given_modern(rest)
        if given is None:
            continue
        if not place_ok(place):
            continue
        # 单字名：家乡须已有地名信号（place_ok），且优先更长家乡
        score = len(place) * 3
        if GEO_TAIL.search(place):
            score += 30
        # 姓+1 / 姓+2 优先于 姓+3（避免 田心刘金 这类）
        if len(rest) == 2:
            score += 50
        elif len(rest) == 3:
            score += 45
        elif len(rest) == 4:
            score += 5
        if GEO_IN_GIVEN.search(given):
            score -= 100
        candidates.append((score, place, rest))

    if not candidates:
        return None
    candidates.sort(key=lambda x: x[0], reverse=True)
    _, place, rest = candidates[0]
    return {"hometown": place, "name": rest, "rule": "modern"}


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--write", action="store_true")
    parser.add_argument("--source", default=str(SOURCE))
    args = parser.parse_args()
    source = Path(args.source)

    if not source.exists():
        raise SystemExit(f"文件不存在: {source}")

    wb = load_workbook(source)
    ws = wb.active
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    idx = {h: i + 1 for i, h in enumerate(headers) if h}
    if "姓名" not in idx or "家乡" not in idx:
        raise SystemExit(f"缺少列: {headers}")

    col_name = idx["姓名"]
    col_home = idx["家乡"]

    changes = []
    skipped = []

    for r in range(2, ws.max_row + 1):
        raw_name = ws.cell(r, col_name).value
        raw_home = ws.cell(r, col_home).value
        name = str(raw_name).strip() if raw_name is not None else ""
        home = str(raw_home).strip() if raw_home is not None else ""
        if not name:
            continue

        # 已是短名且无拆分空间
        if PURE_SHI.fullmatch(name) or (len(name) <= 3 and "氏" not in name):
            continue

        sp = try_split(name)
        if not sp:
            if len(name) >= 4:
                skipped.append({"row": r, "name": name, "home": home, "why": "no-match"})
            continue

        new_home = home if home else sp["hometown"]
        new_name = sp["name"]
        if new_name == name and (home or "") == (new_home or ""):
            continue

        changes.append(
            {
                "row": r,
                "oldName": name,
                "newName": new_name,
                "oldHome": home,
                "newHome": new_home,
                "rule": sp["rule"],
                "homeOverwritten": False,
                "homeFilled": not bool(home),
            }
        )

        if args.write:
            ws.cell(r, col_name).value = new_name
            if not home:
                ws.cell(r, col_home).value = sp["hometown"]

    report = {
        "source": str(source),
        "write": args.write,
        "totalRows": ws.max_row - 1,
        "changed": len(changes),
        "skipped": len(skipped),
        "byRule": {
            "shi": sum(1 for c in changes if c["rule"] == "shi"),
            "modern": sum(1 for c in changes if c["rule"] == "modern"),
        },
        "changes": changes,
        "skippedSamples": skipped[:80],
        "allSkipped": skipped,
    }
    REPORT.parent.mkdir(parents=True, exist_ok=True)
    REPORT.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")

    print(
        json.dumps(
            {
                "source": str(source),
                "write": args.write,
                "changed": len(changes),
                "skipped": len(skipped),
                "byRule": report["byRule"],
                "report": str(REPORT),
            },
            ensure_ascii=False,
            indent=2,
        )
    )

    # 关键样例自检
    tests = [
        "前溪左氏细娘",
        "永新左坊村左积英",
        "上城刘青莲",
        "上城刘龙英",
        "洲湖廖家肖氏秋娘",
        "金田罗坂刘转",
        "刘氏五娘",
        "李雪英",
    ]
    print("self-check:")
    for t in tests:
        print(f"  {t} -> {try_split(t)}")

    if not args.write:
        print("Dry-run only. Re-run with --write to save.")
        return

    stamp = datetime.now().strftime("%Y%m%d%H%M%S")
    backup = source.with_name(f"妻子表_姓名家乡拆分_{stamp}.xlsx")
    wb.save(source)
    shutil.copy2(source, backup)
    print(f"Written: {source}")
    print(f"Backup: {backup}")


if __name__ == "__main__":
    main()
